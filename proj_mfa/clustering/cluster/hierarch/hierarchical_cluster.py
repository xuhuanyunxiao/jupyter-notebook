#!/usr/bin/env python
# -*- coding: utf-8 -*-

import operator
import random
from multiprocessing import cpu_count
import sys
from dao.mysql.spokesman_keyword import SpokesmanKeyword
from utils import java_hashcode
import scipy.cluster.hierarchy as sch
from cluster.feature import nltk_multipro_new
from cluster.feature import tfidfWeight
from entity.ClusterObj import ClusterObj
from entity.cluster_result import SubClusterResult, MergerClusterResult
from utils import unionfind
from utils.logger import logger
import uuid
from dao.mysql.custom_dic import CustomDic

""""
2018-07-11更新:

更改聚类步骤:
1: 标题
2: 内容
3: 合并
"""

# 加载自定义词典
custom_dic = CustomDic()
custom_dicts = custom_dic.get_dicts_en()


def hierarchy_cluster(tfidf_matrix, t):

    # 1. 层次聚类
    # 生成点与点之间的距离矩阵,这里用的cos距离
    disMat = sch.distance.pdist(tfidf_matrix, 'cosine')

    # 进行层次聚类:
    # Z = sch.linkage(disMat)
    Z = sch.linkage(disMat, method='average')

    # 将层级聚类结果以树状图表示出来并保存为plot_dendrogram.png
    # P=sch.dendrogram(Z)
    # plt.savefig('plot_dendrogram.png')

    # 根据linkage matrix Z得到聚类结果:
    labels = sch.fcluster(Z, t=t, criterion='distance')
    return labels


def second_get_return_data(labels, second_corpus):
    label_corpus = zip(labels, second_corpus)

    # 去除-1标签
    filter_label_corpus = filter(lambda (x, y): x != -1, label_corpus)

    # 按label分组,统计每个组出现次数最多的title,统计每个组的时间范围
    purity_group = {}
    for label_corpu in filter_label_corpus:
        label = label_corpu[0]
        first_cluster_id = label_corpu[1][0]
        first_cluster_topic = label_corpu[1][1]

        if label not in purity_group.keys():
            purity_group[label] = {}

        if first_cluster_id not in purity_group[label].keys():
            purity_group[label][first_cluster_id] = []

        purity_group[label][first_cluster_id].append((first_cluster_topic))

    return purity_group


# 媒体排序列表, 优先级依次展示
pre_site_list = ['路透社', '美联社', '法新社', '纽约时报', '华盛顿邮报', '华尔街日报', '金融时报', '彭博社']


def title_comparator(a, b):
    a_site = a[0]
    b_site = b[0]

    a_index = b_index = len(pre_site_list)
    if a_site in pre_site_list:
        a_index = pre_site_list.index(a_site)

    if b_site in pre_site_list:
        b_index = pre_site_list.index(b_site)

    return cmp(a_index, b_index)


def get_sub_cluster_topic(site_list, title_list):
    # 按照媒体优先级 获取标题
    if len(set(site_list) & set(pre_site_list)) != 0:
        zipped_site_title = zip(site_list, title_list)

        sorted_titles = sorted(zipped_site_title, cmp=title_comparator)
        cluster_topic = sorted_titles[0][1]
    else:
        title_map = {}
        for title in title_list:
            if title not in title_map.keys():
                title_map[title] = 1
            else:
                title_map[title] += 1

        sorted_titles = sorted(title_map.iteritems(), key=operator.itemgetter(1), reverse=True)
        cluster_topic = sorted_titles[0][0]
    return cluster_topic


def get_sub_cluster_keywords(title_list):
    word_freq_head = []
    key_word_str = ''

    # 获取关键词, TOP(5)
    if len(title_list) != 0:
        dicts = custom_dicts
        title_cut_list = nltk_multipro_new.cut_words(title_list, dicts)
        filtered_list = filter(lambda x: x != '', title_cut_list)

        # 过滤
        if len(filtered_list) != 0:
            word_frequency_list = tfidfWeight.get_word_frequency(title_cut_list)

            if len(word_frequency_list) >= 6:
                word_freq_head = word_frequency_list[:6]
            else:
                word_freq_head = word_frequency_list

            key_words = [word_freq[0] for word_freq in word_frequency_list][:3]
            key_word_str = '/'.join(key_words)
            key_word_str = '(' + key_word_str + ')'

    # 组合
    return key_word_str, word_freq_head



def get_cluster_topic(site_list, title_list, word_frequency_list):

    # 按照媒体优先级 获取标题
    if len(set(site_list) & set(pre_site_list)) != 0:
        zipped_site_title = zip(site_list, title_list)

        sorted_titles = sorted(zipped_site_title, cmp=title_comparator)
        cluster_topic = sorted_titles[0][1]
    else:
        title_map = {}
        for title in title_list:
            if title not in title_map.keys():
                title_map[title] = 1
            else:
                title_map[title] += 1

        sorted_titles = sorted(title_map.iteritems(), key=operator.itemgetter(1), reverse=True)
        cluster_topic = sorted_titles[0][0]

    # 获取关键词, TOP(5)
    word_freq_dict = {}
    for word_frequency in word_frequency_list:
        word = word_frequency[0]
        frequency = word_frequency[1]
        if word not in word_freq_dict:
            word_freq_dict[word] = 0
        word_freq_dict[word] = word_freq_dict[word] + frequency

    word_freq_dict = sorted(word_freq_dict.items(), key=lambda x: x[1], reverse=True)
    key_words = [word_freq[0] for word_freq in word_freq_dict][:5]
    key_word_str = '/'.join(key_words)
    key_word_str = '(' + key_word_str + ')'

    # 组合
    cluster_topic = cluster_topic + key_word_str
    return cluster_topic


def first_get_return_data(labels, corpus):
    label_corpus = zip(labels, corpus)

    # 去除-1标签
    filter_label_corpus = label_corpus

    # 按label分组, 统计每个组内出现的最长文章
    purity_group = {}
    for line in filter_label_corpus:
        label = line[0]
        id = line[1].messageId
        title = line[1].messageTitle
        content = line[1].messageContent
        publish_time = line[1].messagePublishtime
        site_name = line[1].site_name

        if label not in purity_group.keys():
            purity_group[label] = {}
            purity_group[label]["ids"] = []
            purity_group[label]["title_list"] = []
            purity_group[label]["site_names"] = []
            purity_group[label]["publish_times"] = []
            purity_group[label]["content_list"] = []

        purity_group[label]["ids"].append(id)
        purity_group[label]["title_list"].append(title)
        purity_group[label]["site_names"].append(site_name)
        purity_group[label]["publish_times"].append(publish_time)
        purity_group[label]["content_list"].append(content)

    # 组合title, publish_time 字段
    cluster_result = {}
    for label, value in purity_group.items():
        # ids
        ids = map(str, purity_group[label]["ids"])

        # site_names
        site_names = purity_group[label]["site_names"]

        # title_list
        title_list = purity_group[label]["title_list"]

        # 对title按照指定的规则进行排序
        cluster_topic = get_sub_cluster_topic(site_names, title_list)
        key_word_str, word_freq_head = get_sub_cluster_keywords(title_list)
        cluster_topic = cluster_topic + key_word_str

        # content_list
        content_list = purity_group[label]["content_list"]

        # publish_times
        publish_times = purity_group[label]["publish_times"]

        # 构造 cluster_result
        cluster_result[label] = {}
        cluster_result[label]["ids"] = ids
        cluster_result[label]["site_names"] = site_names
        cluster_result[label]["cluster_topic"] = cluster_topic
        cluster_result[label]["publish_times"] = publish_times
        cluster_result[label]["title_list"] = title_list
        cluster_result[label]["content_list"] = content_list
        cluster_result[label]["word_freq_head"] = word_freq_head

    return cluster_result


def run_cluster(tfidf_matrix, t):
    """
    层次聚类
    :param tfidf_matrix:
    :return:
    """
    label = hierarchy_cluster(tfidf_matrix, t)
    return label


def get_result_corpus(corpus):
    """
    当first_cluster_results 为空, 随机选择10条数据作为聚类结果返回.
    :param corpus:
    :return:
    """
    random_corpus = []
    if len(corpus) > 10:
        random_corpus = random.sample(corpus, 10)  # 从list中随机获取10个元素
    else:
        random_corpus = corpus

    cObjArray = []
    for idx, cluster_message_obj in enumerate(random_corpus):
        # clusterId
        clusterId = idx

        # clusterTopic  对每一title降序排序
        clusterTopic = cluster_message_obj.messageTitle

        # clusterPublishtimeRange 对时间排序
        clusterPublishBeginTime = cluster_message_obj.messagePublishtime
        clusterPublishEndTime = cluster_message_obj.messagePublishtime

        # clusterMember
        clusterMember = str(cluster_message_obj.messageId)

        # cluserMemeberCount
        cluserMemeberCount = 1

        # siteCount
        siteCount = 1

        # 构造ClusterObj
        cObj = ClusterObj(clusterId=clusterId, clusterTopic=clusterTopic,
                          clusterPublishBeginTime=clusterPublishBeginTime, clusterPublishEndTime=clusterPublishEndTime,
                          clusterMember=clusterMember, cluserMemeberCount=cluserMemeberCount,
                          siteCount=siteCount)
        cObjArray.append(cObj)
    return cObjArray


def get_num_processes(corpus_length):
    # 计算num_processes
    num_processes = cpu_count() / 4
    if corpus_length > 30000:
        num_processes = cpu_count()
    elif corpus_length > 10000:
        num_processes = cpu_count() / 2
    num_processes = 4

    return num_processes


def perform_cluster(is_manual, cluster_type, manual_id, subtopic_id, language_type,
                    corpus, min_sample, save_group_id):
    """ 二次聚类 """

    corpus_length = len(corpus)
    if corpus_length == 0:
        logger.info('perform_cluster returning run, because len(corpus) == 0')
        return

    num_processes = get_num_processes(corpus_length)

    """  第一步, 根据标题, 内容聚类  """
    first_articles = [(obj.messageTitle, obj.messageContent) for obj in corpus]
    first_cut_results = nltk_multipro_new.multi_cut_words(first_articles,
                                                          num_processes)
    logger.info('first_cut_results end....')

    ziped_first_corpus_cut_results = zip(corpus, first_cut_results)
    ziped_first_corpus_cut_results = filter(lambda t: t[1] != '', ziped_first_corpus_cut_results)
    corpus = [t[0] for t in ziped_first_corpus_cut_results]
    first_cut_results = [t[1] for t in ziped_first_corpus_cut_results]

    # 构建TF-IDF矩阵
    from cluster.feature import tfidfWeight
    first_tf_matrix = tfidfWeight.getTfidf(first_cut_results, True)
    logger.info('first_tfidf_matrix end....')

    # 聚类算法
    first_labels = run_cluster(first_tf_matrix.toarray(), 0.70)
    first_cluster_results = first_get_return_data(first_labels, corpus)
    logger.info('first_cluster_results end....')

    for label, dict_ in first_cluster_results.items():
        logger.debug('label:%d, cluster_topic: %s, ids: %s' % (label, dict_['cluster_topic'], '-'.join(dict_['ids'])))

    # 如果第一次聚类结果为null
    cluster_results = []
    if len(first_cluster_results) == 0:
        logger.info('first_cluster_results 为空, 随机选择10条数据作为聚类结果....')
        cluster_results = get_result_corpus(corpus)

    else:
        # 二次聚类, 根据topic
        second_corpus = [(label, topic['title_list']) for (label, topic) in first_cluster_results.items()]
        second_articles = [(' '.join(article[1]), '') for article in second_corpus]

        second_cut_results = nltk_multipro_new.multi_cut_words(second_articles, num_processes)
        logger.info('second_cut_results end....')

        ziped_second_corpus_cut_results = zip(second_corpus, second_cut_results)
        ziped_second_corpus_cut_results = filter(lambda t: t[1] != '', ziped_second_corpus_cut_results)
        second_corpus = [t[0] for t in ziped_second_corpus_cut_results]
        second_cut_results = [t[1] for t in ziped_second_corpus_cut_results]

        # 构建TF-IDF矩阵,  减少文章数对TF的影响
        second_tf_matrix_tmp = tfidfWeight.getTfidf(second_cut_results, True)
        second_tf_matrix = []
        for idx, arr in enumerate(second_tf_matrix_tmp.toarray()):
            article_len = len(second_corpus[idx][1])
            arr = [ele / float(article_len) for ele in arr]
            second_tf_matrix.append(arr)
        logger.info('second_tf_matrix end....')

        second_labels = run_cluster(second_tf_matrix, 0.75)
        second_cluster_results = second_get_return_data(second_labels, second_corpus)
        logger.info('second_cluster_results end....')

        # 合并聚类结果
        first_ids = [[label] for (label, topic) in first_cluster_results.items()]
        logger.debug('first_ids: %s', first_ids)

        second_ids = [clusters.keys() for (label, clusters) in second_cluster_results.items() if (len(clusters)) > 1]
        logger.debug('second_ids: %s', second_ids)

        first_ids.extend(second_ids)
        u = unionfind.UnionFind(first_ids)
        u.create_tree()
        ids_list = u.get_tree()
        logger.debug('ids_list: %s', ids_list)

        cluster_results = return_data(ids_list, first_cluster_results, min_sample, is_manual,
                                      cluster_type, manual_id, subtopic_id, language_type, save_group_id)
        logger.info('cluster_results....')
        # test_save_xlsx(ids_list, first_cluster_results)

    return cluster_results


def test_save_xlsx(union_index_list, first_cluster_results):
    data = []
    for idx, cluster_index in enumerate(union_index_list):
        label = str(idx)

        for index in cluster_index:

            title_list = first_cluster_results[index]['title_list']
            content_list = first_cluster_results[index]['content_list']
            zipped_title_content = zip(title_list, content_list)

            for title, content in zipped_title_content:
                data.append([label, title, content])

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    for idx in range(0, len(data)):
        row = idx + 1
        ws.cell(row=row + 1, column=1).value = data[idx][0]
        ws.cell(row=row + 1, column=2).value = data[idx][1]
        ws.cell(row=row + 1, column=3).value = data[idx][2]

    test_file = 'd:\\test.xlsx'
    wb.save(test_file)
    print '保存中间数据完成, file: %s' % test_file


def return_data(cluster_ids_list, first_cluster_results, min_sample, is_manual,
                     cluster_type, manual_id, subtopic_id, language_type, save_group_id):
    ret = []
    for idx, cluster_ids in enumerate(cluster_ids_list):
        ids = [id for cluster_id in cluster_ids for id in first_cluster_results[cluster_id]["ids"]]
        if len(ids) < min_sample:
            continue

        cluster_result_id = get_cluster_result_id(is_manual, cluster_type, manual_id, subtopic_id, save_group_id, idx)

        word_freq_heads = []
        sitenames = []
        publishtimes = []
        sub_cluster_list = []
        for cluster_id in cluster_ids:
            _ids = map(str, first_cluster_results[cluster_id]["ids"])
            _sitenames = first_cluster_results[cluster_id]["site_names"]
            _publishtimes = first_cluster_results[cluster_id]["publish_times"]
            _cluster_topic = first_cluster_results[cluster_id]["cluster_topic"]
            _word_freq_head = first_cluster_results[cluster_id]["word_freq_head"]

            # sub_cluster_result
            sub_cluster = parse_sub_cluster(cluster_id, _cluster_topic, _ids, _sitenames, _publishtimes,
                                            cluster_result_id, cluster_type, language_type, save_group_id,
                                            is_manual, manual_id, subtopic_id)
            sub_cluster_list.append(sub_cluster)

            # 添加
            word_freq_heads.extend(_word_freq_head)
            sitenames.extend(_sitenames)
            publishtimes.extend(_publishtimes)

        # 构造cluster
        cluster = parse_cluster(cluster_result_id, idx, ids, word_freq_heads, sitenames, publishtimes, cluster_type,
                      language_type, save_group_id, is_manual, manual_id, subtopic_id, sub_cluster_list)

        # 添加sub_cluster
        cluster.set_sub_clusters(sub_cluster_list)
        ret.append(cluster)

    # 增加排序, 根据关键词
    sort_bywords(ret)

    return ret


def sort_bywords(list):
    """ 根据关键词对类别进行排序 """
    results = SpokesmanKeyword().get_words()

    for idx, cluster in enumerate(list):
        topic_name = cluster.topic
        keyword_id = -1
        priority = sys.maxint
        for result in results:
            keyword = result['keyword']
            if keyword in topic_name:
                keyword_id = result['id']
                priority = result['priority']
                break
        cluster.keyword_id = keyword_id
        cluster.priority = priority

    # 排序
    list.sort(cmp=sort_cmp)

    # 赋值order_id
    for idx, cluster in enumerate(list):
        cluster.order_id = idx + 1


def sort_cmp(a, b):
    if a.priority == b.priority:
        if a.site_count == b.site_count:
            return int(b.member_count - a.member_count)
        else:
            return int(b.site_count - a.site_count)
    else:
        return int(a.priority - b.priority)


def sub_cluster_comparator(cluster_a, cluster_b):
     member_count_a = cluster_a.cluster_member_count
     member_count_b = cluster_b.cluster_member_count
     if member_count_a != member_count_b:
         return member_count_a > member_count_b

     site_count_a = cluster_a.site_count
     site_count_b = cluster_b.site_count
     return site_count_a > site_count_b


def parse_sub_cluster(cluster_id, cluster_topic, ids, sitenames, publishtimes, cluster_result_id, cluster_type,
                      language_type, group_id, is_manual, manual_id, subtopic_id):
    # _ids 发布时间倒叙排序
    _zipped_id_publistimes = zip(ids, publishtimes)
    _zipped_id_publistimes.sort(key=lambda t: t[1], reverse=True)
    ids = [t[0] for t in _zipped_id_publistimes]
    cluster_member = "^A".join(ids)

    # _publishtimes 排序
    publishtimes.sort()
    cluster_begin_time = publishtimes[0]
    cluster_end_time = publishtimes[-1]

    # cluster_member_count
    cluster_member_count = len(ids)

    # site_count
    site_count = len(set(sitenames))

    obj = SubClusterResult(cluster_id, cluster_topic,
                           cluster_begin_time, cluster_end_time,
                           cluster_member, cluster_member_count,
                           site_count, cluster_result_id, cluster_type,
                           language_type, group_id,
                           is_manual, manual_id, subtopic_id)

    return obj


def parse_cluster(cluster_result_id, cluster_id, ids, word_freq_heads, sitenames, publishtimes,
                  cluster_type, language_type, group_id, is_manual, manual_id, subtopic_id,
                  sub_cluster_list):
    # cluster_topic
    sub_cluster_list = sorted(sub_cluster_list, key=lambda x: x.member_count, reverse=True)
    cluster_topic = sub_cluster_list[0].topic

    index = cluster_topic.find('(')
    if index != -1:
        cluster_topic = cluster_topic[:index]

    # word_freq_heads
    word_freq_map = {}
    for word_freq_head in word_freq_heads:
        word = word_freq_head[0]
        freq = word_freq_head[1]
        if word not in word_freq_map:
            word_freq_map[word] = 0
        word_freq_map[word] = word_freq_map[word] + freq

    sorted_words = sorted(word_freq_map.iteritems(), key=operator.itemgetter(1), reverse=True)
    keywords = [sorted_word[0] for sorted_word in sorted_words][:6]
    keyword_str = '/'.join(keywords)
    keyword_str = '(' + keyword_str + ')'
    cluster_topic = cluster_topic + keyword_str

    # publishtimes排序
    publishtimes.sort()  # 时间升序
    cluster_begin_time = publishtimes[0]
    cluster_end_time = publishtimes[-1]

    # member
    member = '^A'.join(ids)

    # cluster_member_count
    cluster_member_count = len(ids)

    # site_count
    site_count = len(set(sitenames))

    # 创建MergerClusterResult
    obj = MergerClusterResult(cluster_result_id, cluster_id, cluster_topic,
                              cluster_begin_time,
                              cluster_end_time,
                              member, cluster_member_count, site_count,
                              cluster_type, language_type, group_id,
                              is_manual, manual_id, subtopic_id)
    return obj


def get_cluster_result_id(is_manual, cluster_type, manual_id, subtopic_id, save_group_id, id):
    """ 获得id """
    name = uuid.uuid1()
    hashcode = java_hashcode.get_hashcode(str(name))
    no_negative_hashcode = abs(hashcode)
    return no_negative_hashcode
