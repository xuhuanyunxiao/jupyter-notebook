#!/usr/bin/python
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from entity import ClusterMessageObj
from db_mysql import BaseMysql


class TestBaseDataView(BaseMysql):
    def __init__(self):
        super(TestBaseDataView, self).__init__()
        self.table_name = 'test_base_data_view'

    @staticmethod
    def read_from_excel(file_path):
        """
        从excel读取数据
        """
        test_corpus = []
        workbook = load_workbook(file_path)
        sheet_names = workbook.sheetnames  # 获得表单名字
        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            for row in range(1, sheet.max_row + 1):
                row = row
                title_column = 2
                content_column = 3
                title = sheet.cell(row=row, column=title_column).value
                content = sheet.cell(row=row, column=content_column).value
                test_corpus.append((title, content))
        workbook.close()
        return test_corpus

    def save(self):
        """
        :param corpus:
        :return:
        """
        corpus = TestBaseDataView.read_from_excel('D:\\1000.xlsx')
        values = []
        for idx, corpu in enumerate(corpus):
            value = []
            value.append(idx + 1)
            value.append(str(corpu[0]).encode('utf-8'))
            value.append(str(corpu[1]).encode('utf-8'))
            value.append('2018-07-11 17:00:00')
            value.append('test')

            values.append(value)

        table = self.table_name
        attrs = ['id', 'title', 'content', 'publishtime', 'site_name']
        self.insert_many(table, attrs, values)

    def get_corpus(self):
        """
        获得聚类语料
        :param start_time:
        :param end_time:
        :param group_id:
        :return:
        """
        sql = """
        SELECT id, title, content, publishtime, site_name 
        FROM test_base_data_view 
       """

        records = self.fetch_all(sql)

        return TestBaseDataView.parse_corpus_records(records)

    @staticmethod
    def parse_corpus_records(records):
        data = []
        for row in records:
            id = row['id']

            title = row['title']
            title = title.encode('utf-8')

            content = row['content']
            content = content.encode('utf-8')

            publish_time = row['publishtime'].strftime("%Y-%m-%d %H:%M:%S").decode('utf-8')
            site_name = row['site_name']
            site_name = site_name.encode('utf-8')

            cluster_obj = ClusterMessageObj(id, title, publish_time, content, site_name)
            data.append(cluster_obj)
        return data
