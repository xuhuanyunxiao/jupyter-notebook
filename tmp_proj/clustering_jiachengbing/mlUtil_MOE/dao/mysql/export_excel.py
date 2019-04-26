#!/usr/bin/python
# -*- coding: utf-8 -*-

from db_mysql import BaseMysql
from openpyxl import Workbook


class SpokesmanKeyword(BaseMysql):
    def __init__(self):
        super(SpokesmanKeyword, self).__init__()
        self.table_name = ''

    def get_corpus(self):
        sql = """
        SELECT id, title, content 
        FROM base_data_view 
        WHERE publishtime >= '2018-10-10 09:08:00' AND publishtime < '2018-10-10 12:01:00' 
        AND language_type = 1 
        AND group_id IN (7, 16)
        group by titlehash
        """

        records = self.fetch_all(sql)
        return records

    def get_topics(self, manual_id):
        sql = """
        SELECT b.cluster_id as sub_cluster_id, b.cluster_topic sub_cluster_topic, 
        b.cluster_member cluster_member, a.cluster_id cluster_id, a.cluster_topic cluster_topic 
        FROM cluster_result a, sub_cluster_result b 
        WHERE a.manual_id = %d AND b.cluster_result_id = a.id
        """ % (manual_id,)

        records = self.fetch_all(sql)
        return records

    def get_articles(self, ids):
        ids = [str(id) for id in ids]
        ids_str = ','.join(ids)
        ids_str = '(' + ids_str + ')'

        sql = """
        SELECT title, content 
        FROM base_data_view 
        WHERE id IN %s
        """ % (ids_str,)
        records = self.fetch_all(sql)
        return records

    @staticmethod
    def save_excel(records, filename):
        # 将预处理的完毕的数据保存到中间文件中
        mid_wb = Workbook()
        mid_ws = mid_wb.active
        mid_ws.cell(row=1, column=1).value = 'cluster_id'
        mid_ws.cell(row=1, column=2).value = 'cluster_topic'
        mid_ws.cell(row=1, column=3).value = 'sub_cluster_id'
        mid_ws.cell(row=1, column=4).value = 'sub_cluster_topic'
        mid_ws.cell(row=1, column=5).value = 'title'
        mid_ws.cell(row=1, column=6).value = 'content'
        for row in range(0, len(records)):
            mid_ws.cell(row=row+2, column=1).value = records[row][0]
            mid_ws.cell(row=row+2, column=2).value = records[row][1]
            mid_ws.cell(row=row+2, column=3).value = records[row][2]
            mid_ws.cell(row=row+2, column=4).value = records[row][3]
            mid_ws.cell(row=row+2, column=5).value = records[row][4]
            mid_ws.cell(row=row+2, column=6).value = records[row][5]
        mid_wb.save(filename)
        print('save success: ', filename)

    def export(self, manual_id, filename):
        records = []

        # 聚类
        topics = self.get_topics(manual_id)
        has_ids = []
        for topic in topics:
            cluster_id = topic['cluster_id']
            cluster_topic = topic['cluster_topic']

            sub_cluster_id = topic['sub_cluster_id']
            sub_cluster_topic = topic['sub_cluster_topic']
            cluster_member = topic['cluster_member']

            ids = cluster_member.split('^A')
            has_ids.extend(ids)
            articles = self.get_articles(ids)
            for article in articles:
                title = article['title']
                content = article['content']

                records.append((cluster_id, cluster_topic, sub_cluster_id, sub_cluster_topic, title, content))

        # 未聚类数据
        corpus = self.get_corpus()
        for corpu in corpus:
            id = corpu['id']
            title = corpu['title']
            content = corpu['content']
            if str(id) not in has_ids:
                records.append((-1, '', -1, '', title, content))

        SpokesmanKeyword.save_excel(records, filename)


if __name__ == '__main__':

    spokesmanKeyword = SpokesmanKeyword()
    spokesmanKeyword.export(73, 'save.xlsx')



