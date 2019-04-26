#!/usr/bin/env python
# -*- coding: utf-8 -*-

import operator
import random
from multiprocessing import cpu_count
from utils import java_hashcode
import scipy.cluster.hierarchy as sch
from cluster.feature import nltk_multipro_new
from entity.ClusterObj import ClusterObj
from entity.cluster_result import SubClusterResult, MergerClusterResult
from utils import unionfind
from utils.logger import logger
import uuid

""""
2018-07-11更新:

更改聚类步骤:
1: 标题
2: 内容
3: 合并
"""


def hierarchy_cluster(tfidf_matrix):
    # 1. 层次聚类
    # 生成点与点之间的距离矩阵,这里用的cos距离
    disMat = sch.distance.pdist(tfidf_matrix, 'cosine')

    # 进行层次聚类:
    # Z = sch.linkage(disMat)
    Z = sch.linkage(disMat, method='average')

    # 将层级聚类结果以树状图表示出来并保存为plot_dendrogram.png
    # P=sch.dendrogram(Z)
    # plt.savefig('plot_dendrogram.png')

    # 根据linkage matrix Z得到聚类结果:
    labels = sch.fcluster(Z, t=0.70, criterion='distance')
    return labels


def second_get_return_data(labels):
    # 按label分组, 统计每个组内出现的最长文章
    purity_group = {}
    for idx, label in enumerate(labels):
        if label not in purity_group.keys():
            purity_group[label] = []

        purity_group[label].append(idx)

    return purity_group


# 媒体排序列表, 优先级一次展示
pre_site_list = ['路透社', '美联社', '法新社', '纽约时报', '华盛顿邮报', '华尔街日报', '金融时报', '彭博社']


def title_comparator(a, b):
    a_site = a[0]
    b_site = b[0]

    a_index = b_index = len(pre_site_list)
    if a_site in pre_site_list:
        a_index = pre_site_list.index(a_site)

    if b_site in pre_site_list:
        b_index = pre_site_list.index(b_site)

    return a_index > b_index


def get_cluster_topic(site_list, title_list):
    if len(set(site_list) & set(pre_site_list)) != 0:
        zipped_site_title = zip(site_list, title_list)

        sorted_titles = sorted(zipped_site_title, cmp=title_comparator)
        cluster_topic = sorted_titles[0][1]
    else:
        title_map = {}
        for title in title_list:
            if title not in title_map.keys():
                title_map[title] = 1
            else:
                title_map[title] += 1

        sorted_titles = sorted(title_map.iteritems(), key=operator.itemgetter(1), reverse=True)
        cluster_topic = sorted_titles[0][0]
    return cluster_topic


def first_get_return_data(labels):
    # 按label分组, 统计每个组内出现的最长文章
    purity_group = {}
    for idx, label in enumerate(labels):
        if label not in purity_group.keys():
            purity_group[label] = []

        purity_group[label].append(idx)

    return purity_group


def run_cluster(tfidf_matrix):
    """
    层次聚类
    :param tfidf_matrix:
    :return:
    """
    label = hierarchy_cluster(tfidf_matrix)
    return label


def get_result_corpus(corpus):
    """
    当first_cluster_results 为空, 随机选择10条数据作为聚类结果返回.
    :param corpus:
    :return:
    """
    random_corpus = []
    if len(corpus) > 10:
        random_corpus = random.sample(corpus, 10)  # 从list中随机获取10个元素
    else:
        random_corpus = corpus

    cObjArray = []
    for idx, cluster_message_obj in enumerate(random_corpus):
        # clusterId
        clusterId = idx

        # clusterTopic  对每一title降序排序
        clusterTopic = cluster_message_obj.messageTitle

        # clusterPublishtimeRange 对时间排序
        clusterPublishBeginTime = cluster_message_obj.messagePublishtime
        clusterPublishEndTime = cluster_message_obj.messagePublishtime

        # clusterMember
        clusterMember = str(cluster_message_obj.messageId)

        # cluserMemeberCount
        cluserMemeberCount = 1

        # siteCount
        siteCount = 1

        # 构造ClusterObj
        cObj = ClusterObj(clusterId=clusterId, clusterTopic=clusterTopic,
                          clusterPublishBeginTime=clusterPublishBeginTime, clusterPublishEndTime=clusterPublishEndTime,
                          clusterMember=clusterMember, cluserMemeberCount=cluserMemeberCount,
                          siteCount=siteCount)
        cObjArray.append(cObj)
    return cObjArray


def get_num_processes(corpus_length):
    # 计算num_processes
    num_processes = cpu_count() / 4
    if corpus_length > 30000:
        num_processes = cpu_count()
    elif corpus_length > 10000:
        num_processes = cpu_count() / 2
    num_processes = 4

    return num_processes


def perform_cluster(is_manual, cluster_type, manual_id, subtopic_id, language_type,
                    corpus, min_sample, save_group_id):
    """ 二次聚类 """

    corpus_length = len(corpus)
    if corpus_length == 0:
        logger.info('perform_cluster returning run, because len(corpus) == 0')
        return

    num_processes = get_num_processes(corpus_length)

    """  第一步, 根据标题聚类  """
    first_articles = [(obj.messageTitle, '') for obj in corpus]
    first_cut_results = nltk_multipro_new.multi_cut_words(first_articles,
                                                          num_processes)
    logger.info('first_cut_results end....')

    # 构建TF-IDF矩阵
    from cluster.feature import tfidfWeight
    first_tf_matrix = tfidfWeight.getTfidf(first_cut_results, True)
    logger.info('first_tfidf_matrix end....')

    # 聚类算法
    first_labels = run_cluster(first_tf_matrix.toarray())
    first_cluster_results = first_get_return_data(first_labels)
    logger.info('first_cluster_results end....')

    for label, dict_ in first_cluster_results.items():
        logger.debug('label:%d, cluster_topic: %s, ids: %s' % (label, dict_['cluster_topic'], '-'.join(dict_['ids'])))

    """  第一步, 根据内容聚类  """
    second_articles = [('', obj.messageContent) for obj in corpus]
    second_cut_results = nltk_multipro_new.multi_cut_words(second_articles, num_processes)
    logger.info('second_cut_results end....')

    # 构建TF-IDF矩阵,  减少文章数对TF的影响
    second_tf_matrix = tfidfWeight.getTfidf(second_cut_results, True)
    logger.info('second_tf_matrix end....')

    second_labels = run_cluster(second_tf_matrix)
    second_cluster_results = second_get_return_data(second_labels)
    logger.info('second_cluster_results end....')

    # 合并聚类结果
    first_ids = first_cluster_results.values()
    second_ids = second_cluster_results.values()

    first_ids.extend(second_ids)
    u = unionfind.UnionFind(first_ids)
    u.create_tree()
    ids_list = u.get_tree()
    logger.debug('ids_list: %s', ids_list)

    cluster_results = return_data(ids_list, corpus, min_sample, is_manual,
                                  cluster_type, manual_id, subtopic_id, language_type, save_group_id)
    logger.info('cluster_results....')
    # test_save_xlsx(ids_list, first_cluster_results)

    return cluster_results


def test_save_xlsx(union_index_list, first_cluster_results):
    data = []
    for idx, cluster_index in enumerate(union_index_list):
        label = str(idx)

        for index in cluster_index:

            title_list = first_cluster_results[index]['title_list']
            content_list = first_cluster_results[index]['content_list']
            zipped_title_content = zip(title_list, content_list)

            for title, content in zipped_title_content:
                data.append([label, title, content])

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    for idx in range(0, len(data)):
        row = idx + 1
        ws.cell(row=row + 1, column=1).value = data[idx][0]
        ws.cell(row=row + 1, column=2).value = data[idx][1]
        ws.cell(row=row + 1, column=3).value = data[idx][2]

    test_file = 'd:\\test.xlsx'
    wb.save(test_file)
    print '保存中间数据完成, file: %s' % test_file


def return_data(cluster_ids_list, corpus, min_sample, is_manual,
                     cluster_type, manual_id, subtopic_id, language_type, save_group_id):
    ret = []
    for idx, corpus_idxs in enumerate(cluster_ids_list):
        if len(corpus_idxs) < min_sample:
            continue

        cluster_result_id = get_cluster_result_id(is_manual, cluster_type, manual_id, subtopic_id, save_group_id, idx)

        titles = []
        sitenames = []
        publishtimes = []
        sub_cluster_list = []
        for corpus_idx in corpus_idxs:
            _ids = map(str, corpus[corpus_idx]["ids"])
            # _sitenames = first_cluster_results[cluster_id]["site_names"]
            # _title_list = first_cluster_results[cluster_id]["title_list"]
            # _publishtimes = first_cluster_results[cluster_id]["publish_times"]
            # _cluster_topic = first_cluster_results[cluster_id]["cluster_topic"]

            # # sub_cluster_result
            # sub_cluster = parse_sub_cluster(cluster_id, _cluster_topic, _ids, _sitenames, _publishtimes,
            #                                 cluster_result_id, cluster_type, language_type, save_group_id,
            #                                 is_manual, manual_id, subtopic_id)
            # sub_cluster_list.append(sub_cluster)
            #
            # # 添加
            # titles.extend(_title_list)
            # sitenames.extend(_sitenames)
            # publishtimes.extend(_publishtimes)

        # 构造cluster
        cluster = parse_cluster(cluster_result_id, idx, ids, titles, sitenames, publishtimes, cluster_type,
                      language_type, save_group_id, is_manual, manual_id, subtopic_id)

        # 添加sub_cluster
        cluster.set_sub_clusters(sub_cluster_list)
        ret.append(cluster)

    return ret


def sub_cluster_comparator(cluster_a, cluster_b):
     member_count_a = cluster_a.cluster_member_count
     member_count_b = cluster_b.cluster_member_count
     if member_count_a != member_count_b:
         return member_count_a > member_count_b

     site_count_a = cluster_a.site_count
     site_count_b = cluster_b.site_count
     return site_count_a > site_count_b


def parse_sub_cluster(cluster_id, cluster_topic, ids, sitenames, publishtimes, cluster_result_id, cluster_type,
                      language_type, group_id, is_manual, manual_id, subtopic_id):
    # _ids 发布时间倒叙排序
    _zipped_id_publistimes = zip(ids, publishtimes)
    _zipped_id_publistimes.sort(key=lambda t: t[1], reverse=True)
    ids = [t[0] for t in _zipped_id_publistimes]
    cluster_member = "^A".join(ids)

    # _publishtimes 排序
    publishtimes.sort()
    cluster_begin_time = publishtimes[0]
    cluster_end_time = publishtimes[-1]

    # cluster_member_count
    cluster_member_count = len(ids)

    # site_count
    site_count = len(set(sitenames))

    obj = SubClusterResult(cluster_id, cluster_topic,
                           cluster_begin_time, cluster_end_time,
                           cluster_member, cluster_member_count,
                           site_count, cluster_result_id, cluster_type,
                           language_type, group_id,
                           is_manual, manual_id, subtopic_id)

    return obj


def parse_cluster(cluster_result_id, cluster_id, ids, titles, sitenames, publishtimes,
                  cluster_type, language_type, group_id, is_manual, manual_id, subtopic_id):
    # cluster_topic
    cluster_topic = get_cluster_topic(sitenames, titles)

    # publishtimes排序
    publishtimes.sort()  # 时间升序
    cluster_begin_time = publishtimes[0]
    cluster_end_time = publishtimes[-1]

    # member
    member = '^A'.join(ids)

    # cluster_member_count
    cluster_member_count = len(ids)

    # site_count
    site_count = len(set(sitenames))

    obj = MergerClusterResult(cluster_result_id, cluster_id, cluster_topic,
                              cluster_begin_time,
                              cluster_end_time,
                              member, cluster_member_count, site_count,
                              cluster_type, language_type, group_id,
                              is_manual, manual_id, subtopic_id)
    return obj


def get_cluster_result_id(is_manual, cluster_type, manual_id, subtopic_id, save_group_id, id):
    """ 获得id """
    name = uuid.uuid1()
    hashcode = java_hashcode.get_hashcode(str(name))
    no_negative_hashcode = abs(hashcode)
    return no_negative_hashcode
