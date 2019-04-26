#!/usr/bin/env python
# -*- coding: utf-8 -*-

import operator
import random
from multiprocessing import cpu_count

import scipy.cluster.hierarchy as sch

from cluster.feature import nltk_multipro_new
from cluster.feature import tfidfWeight
from entity.ClusterObj import ClusterObj
from utils import unionfind
from utils.logger import logger


def hierarchy_cluster(tfidf_matrix):
    # 1. 层次聚类
    # 生成点与点之间的距离矩阵,这里用的欧氏距离:
    disMat = sch.distance.pdist(tfidf_matrix, 'cosine')

    # 进行层次聚类:
    Z = sch.linkage(disMat, method='average')

    # 将层级聚类结果以树状图表示出来并保存为plot_dendrogram.png
    # P=sch.dendrogram(Z)
    # plt.savefig('plot_dendrogram.png')

    # 根据linkage matrix Z得到聚类结果:
    labels = sch.fcluster(Z, t=0.70, criterion='distance')
    return labels


def second_get_return_data(labels, second_corpus):
    label_corpus = zip(labels, second_corpus)

    # 去除-1标签
    filter_label_corpus = filter(lambda (x, y): x != -1, label_corpus)

    # 按label分组,统计每个组出现次数最多的title,统计每个组的时间范围
    purity_group = {}
    for label_corpu in filter_label_corpus:
        label = label_corpu[0]
        first_cluster_id = label_corpu[1][0]
        first_cluster_topic = label_corpu[1][1]

        if label not in purity_group.keys():
            purity_group[label] = {}

        if first_cluster_id not in purity_group[label].keys():
            purity_group[label][first_cluster_id] = []

        purity_group[label][first_cluster_id].append((first_cluster_topic))

    return purity_group


def parse_labels(labels):
    # 按label分组, 统计每个组内出现的文章下标
    group = {}
    for idx, label in enumerate(labels):
        if label not in group.keys():
            group[label] = []

        group[label].append(idx)

    return group


def run_cluster(tfidf_matrix):
    """
    层次聚类
    :param tfidf_matrix:
    :param mObjArray:
    :return:
    """
    label = hierarchy_cluster(tfidf_matrix.toarray())
    return label


def get_result_corpus(corpus):
    """
    当first_cluster_results 为空, 随机选择10条数据作为聚类结果返回.
    :param corpus:
    :return:
    """
    random_corpus = []
    if len(corpus) > 10:
        random_corpus = random.sample(corpus, 10)  # 从list中随机获取10个元素
    else:
        random_corpus = corpus

    cObjArray = []
    for idx, cluster_message_obj in enumerate(random_corpus):
        # clusterId
        clusterId = idx

        # clusterTopic  对每一title降序排序
        clusterTopic = cluster_message_obj.messageTitle

        # clusterPublishtimeRange 对时间排序
        clusterPublishBeginTime = cluster_message_obj.messagePublishtime
        clusterPublishEndTime = cluster_message_obj.messagePublishtime

        # clusterMember
        clusterMember = str(cluster_message_obj.messageId)

        # cluserMemeberCount
        cluserMemeberCount = 1

        # siteCount
        siteCount = 1

        # 构造ClusterObj
        cObj = ClusterObj(clusterId=clusterId, clusterTopic=clusterTopic,
                          clusterPublishBeginTime=clusterPublishBeginTime, clusterPublishEndTime=clusterPublishEndTime,
                          clusterMember=clusterMember, cluserMemeberCount=cluserMemeberCount,
                          siteCount=siteCount)
        cObjArray.append(cObj)
    return cObjArray


def get_num_processes(corpus_length):
    # 计算num_processes
    num_processes = cpu_count() / 4
    if corpus_length > 30000:
        num_processes = cpu_count()
    elif corpus_length > 10000:
        num_processes = cpu_count() / 2
    num_processes = 2

    return num_processes


def perform_cluster(corpus, min_sample):
    """ 二次聚类 """

    corpus_length = len(corpus)
    if corpus_length == 0:
        logger.info('perform_cluster returning run, because len(corpus) == 0')
        return

    num_processes = get_num_processes(corpus_length)

    """  第一步, 根据标题进行聚类, 将标题相似的文章找出  """
    first_articles = [(obj.messageTitle, '') for obj in corpus]
    first_cut_results = nltk_multipro_new.multi_cut_words(first_articles, num_processes)
    logger.info('first_cut_results end....')

    ziped_first_corpus_cut_results = zip(corpus, first_cut_results)
    ziped_first_corpus_cut_results = filter(lambda t: t[1] != '', ziped_first_corpus_cut_results)
    corpus = [t[0] for t in ziped_first_corpus_cut_results]
    first_cut_results = [t[1] for t in ziped_first_corpus_cut_results]

    # 构建TF-IDF矩阵
    first_tf_matrix = tfidfWeight.getTfidf(first_cut_results, True)
    logger.info('first_tfidf_matrix end....')

    # 聚类算法
    first_labels = run_cluster(first_tf_matrix)
    first_cluster_results = parse_labels(first_labels)
    logger.info('first_cluster_results end....')

    """ 二次聚类, 根据内容进行聚类, 将内容相似的文章找出 """
    second_articles = [('', obj.messageContent) for obj in corpus]
    second_cut_results = nltk_multipro_new.multi_cut_words(second_articles, num_processes)
    logger.info('second_cut_results end....')

    # 构建TF-IDF矩阵
    second_tf_matrix = tfidfWeight.getTfidf(second_cut_results, True)
    logger.info('second_tf_matrix end....')

    # 聚类算法
    second_labels = run_cluster(second_tf_matrix)
    second_cluster_results = parse_labels(second_labels)
    logger.info('second_cluster_results end....')

    """ 合并二次聚类结果 """
    first_index_list = first_cluster_results.values()
    second_index_list = second_cluster_results.values()

    index_list = []
    index_list.extend(first_index_list)
    index_list.extend(second_index_list)

    u = unionfind.UnionFind(index_list)
    u.create_tree()
    union_index_list = u.get_tree()
    logger.info('union_index_list end.')

    """ 返回结果 """
    cluster_results = return_data(union_index_list, corpus, min_sample)
    logger.info('cluster_results....')
    return cluster_results

    # test_save_xlsx(union_index_list, corpus)
    # logger.info('cluster_results....')


def test_save_xlsx(union_index_list, corpus):
    data = []
    for idx, cluster_index in enumerate(union_index_list):
        label = str(idx)

        for index in cluster_index:
            title = corpus[index].messageTitle
            content = corpus[index].messageContent
            data.append([label, title, content])

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    for idx in range(0, len(data)):
        row = idx + 1
        ws.cell(row=row + 1, column=1).value = data[idx][0]
        ws.cell(row=row + 1, column=2).value = data[idx][1]
        ws.cell(row=row + 1, column=3).value = data[idx][2]

    test_file = 'd:\\test.xlsx'
    wb.save(test_file)
    print '保存中间数据完成, file: %s' % test_file


def return_data(union_index_list, corpus, min_sample):
    group = {}
    for idx, cluster_index in enumerate(union_index_list):
        label = idx
        group[label] = {}
        group[label]["ids"] = []
        group[label]["title_list"] = {}
        group[label]["site_names"] = set()
        group[label]["publish_times"] = []

        for index in cluster_index:
            id = corpus[index].messageId
            title = corpus[index].messageTitle
            publish_time = corpus[index].messagePublishtime
            site_name = corpus[index].site_name

            group[label]["ids"].append(id)
            group[label]["site_names"].add(site_name)
            group[label]["publish_times"].append(publish_time)
            if title not in group[label]["title_list"].keys():
                group[label]["title_list"][title] = 1
            else:
                group[label]["title_list"][title] += 1

    cluster_result = []
    for label, value in group.items():
        # clusterId
        cluster_id = label

        # cluster_topic  对每一title降序排序
        titles = group[label]["title_list"]
        sorted_titles = sorted(titles.iteritems(), key=operator.itemgetter(1), reverse=True)
        cluster_topic = sorted_titles[0][0]

        # clusterPublishtimeRange 对时间排序
        publish_times = group[label]["publish_times"]
        publish_times.sort()  # 时间升序
        cluster_publish_beginTime = publish_times[0]
        cluster_publish_endTime = publish_times[-1]

        # cluster_member
        ids_list = map(str, group[label]["ids"])
        cluster_member = "^A".join(ids_list)

        # cluster_member_count
        cluster_member_count = len(ids_list)

        # siteCount
        siteCount = len(group[label]["site_names"])

        # 构造ClusterObj
        if cluster_member_count >= min_sample:
            cObj = ClusterObj(clusterId=cluster_id, clusterTopic=cluster_topic,
                              clusterPublishBeginTime=cluster_publish_beginTime,
                              clusterPublishEndTime=cluster_publish_endTime,
                              clusterMember=cluster_member, cluserMemeberCount=cluster_member_count,
                              siteCount=siteCount)
            cluster_result.append(cObj)
    return cluster_result
